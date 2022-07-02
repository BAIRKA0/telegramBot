from telethon import TelegramClient, events
from telethon.tl.types import InputPhoneContact
from telethon.tl.functions.contacts import ImportContactsRequest
from telethon import functions, types
from openpyxl import load_workbook
import logging
import configparser

logging.basicConfig(format='[%(levelname) 5s/%(asctime)s] %(name)s: %(message)s',level=logging.WARNING)

config = configparser.ConfigParser()
config.read('config.ini')
api_id = int(config['Bot']['api_id'])
api_hash = str(config['Bot']['api_hash'])
channel_id = int(config['Bot']['channel_id'])
owner_id = int(config['Bot']['owner_id'])
path = "телефоны.xlsx"
client = TelegramClient('session_name1', api_id, api_hash)
client.start()

@client.on(events.NewMessage(incoming=True))
async def excel_event_handler(event):
    if event.message.media==None:
        userMessage = event.message.to_dict()['message']
        try:
            sender_id = event.message.from_id.user_id
            print("Сообщение:", str(userMessage), "от пользователя:",sender_id)
        except:
            channel = event.message.peer_id.channel_id
            print("Сообщение:", str(userMessage), "из канала:",channel)

@client.on(events.NewMessage(pattern='/help'))
async def excel_event_handler(event):
    await event.respond('Все команды: \n'
                    '/change - изменение файла номеров (вместе с командой надо отправлять excel файл) '
                    '\n/invite - приглашение пользователей в определенный канал'
                    '\n/run - приглашение пользователей в канал, в котором написали команду'
                    '\n/add - добавление номера в excel (пример: /add 7xxxxxxxxxx)'
                    '\n/del - Удаление номера из excel (пример: /del 7xxxxxxxxxx)'
                    '\n/change_owner - меняет владельца бота и в конце нужно указать пароль (пример /change_owner 7xxxxxxxxxx 1234)')

@client.on(events.NewMessage(from_users=[owner_id] or 'me',pattern='/change_owner'))
async def excel_event_handler(event):
    num = str(event.message.text)
    password = num[(len(num)-4):len(num)]
    if password==config['Bot']['password']:
        num = num.replace('/change_owner ', '')
        num = num.replace(f' {password}','')
        try:
            num = int(num)
            try:
                contact = InputPhoneContact(client_id=0, phone="+" + str(num), first_name=str(num), last_name="")
                result = await client(ImportContactsRequest([contact]))
                id = int(result.imported[0].user_id)
                config.set('Bot', 'owner_id', str(id))
                with open('config.ini', 'w') as configfile:
                    config.write(configfile)
                await event.respond('Владелец изменен')
            except:
                await event.respond('Невозможно добавить контакт')
        except:
            await event.respond('Неправильный номер')
    else:
        await event.respond('Неправильный пароль')

@client.on(events.NewMessage(pattern='/run'))
async def excel_event_handler(event):
    try:
        channel = event.message.peer_id.channel_id
        channel = f'-100{channel}'
        wb = load_workbook(path)
        sheets_name = wb.sheetnames
        sheet = wb[sheets_name[0]]
        r = 1
        z = 199
        col = 0
        while sheet.cell(row=1, column=2).value != sheet.cell(row=r, column=1).value:
            r = r + 1
        while z != 0:
            num = sheet.cell(row=r, column=1).value
            if num != None:
                contact = InputPhoneContact(client_id=0, phone="+" + str(num), first_name=str(num), last_name="")
                result = await client(ImportContactsRequest([contact]))
                id = result.imported[0].user_id
                user = await client.get_entity(types.PeerUser(int(id)))
                try:
                        # await client(InviteToChannelRequest(channel_id,[user]))
                    await client(functions.channels.EditAdminRequest(
                        int(channel),
                        user.id,
                        admin_rights=types.ChatAdminRights(other=True),
                        rank=''
                    ))
                    await client(functions.channels.EditAdminRequest(
                        int(channel),
                        user.id,
                        admin_rights=types.ChatAdminRights(other=False),
                        rank=''
                    ))
                    print(f'{num} Добавлен (id: {id})')
                    col = col + 1
                    r = r + 1
                    z = z - 1
                except:
                    print(f'{num} Не добавлен ')
                    r = r + 1
            else:
                z = 0
        str1 = f'Приглашено пользователей: {col}'
        await event.respond(str1)
    except:
        print('Ошибка')
        await event.respond('Эту команду нужно писать в канале')

@client.on(events.NewMessage(incoming=True, from_users=[owner_id] or 'me', pattern='/add'))
async def excel_event_handler(event):
    num = str(event.message.text)
    num = num.replace('/add ','')
    if len(num)==11 and num[0]=='7':
        wb = load_workbook(path)
        sheets_name = wb.sheetnames
        sheet = wb[sheets_name[0]]
        i=0
        j=1
        while(i!=1):
            if(sheet.cell(row=j,column=1).value!=None):
                j = j + 1
            else:
                sheet.cell(row=j, column=1, value=num)
                i = 1
        wb.save(path)
        await event.respond('Номер добавлен')
    else:
        await event.respond('Неверно указан номер')

@client.on(events.NewMessage(incoming=True, from_users=[owner_id] or 'me', pattern='/del'))
async def excel_event_handler(event):
    num = str(event.message.text)
    num = num.replace('/del ', '')
    if len(num)==11 and num[0]=='7':
        wb = load_workbook(path)
        sheets_name = wb.sheetnames
        sheet = wb[sheets_name[0]]
        i = 0
        for cell in sheet['A']:
            if str(num) == str(cell.value):
                sheet.delete_rows(cell.row)
                wb.save(path)
                i = 1
        if i==1:
            await event.respond('Номер удален')
        else:
            await event.respond('Данного номера нет в списке')
    else:
        await event.respond('Неверно указан номер')

@client.on(events.NewMessage(incoming=True, from_users=[owner_id] or 'me', pattern='/change'))
async def excel_event_handler(event):
    if event.message.media:
        if event.message.file.ext == '.xlsx':
            await event.download_media(file='телефоны.xlsx')

@client.on(events.NewMessage(incoming=True, from_users=[owner_id] or 'me', pattern='/invite'))
async def excel_event_handler(event):
    wb = load_workbook(path)
    sheets_name = wb.sheetnames
    sheet = wb[sheets_name[0]]
    r = 1
    z = 199
    col = 0
    while sheet.cell(row=1, column=2).value != sheet.cell(row=r, column=1).value:
        r=r+1
    while z!=0:
        num = sheet.cell(row=r, column=1).value
        if num!=None:
            try:
                contact = InputPhoneContact(client_id=0, phone="+" + str(num), first_name=str(num), last_name="")
                result = await client(ImportContactsRequest([contact]))
                id = result.imported[0].user_id
                user = await client.get_entity(types.PeerUser(int(id)))
                try:
                    # await client(InviteToChannelRequest(channel_id,[user]))
                    await client(functions.channels.EditAdminRequest(
                        channel_id,
                        user.id,
                        admin_rights=types.ChatAdminRights(other=True),
                        rank=''
                    ))
                    await client(functions.channels.EditAdminRequest(
                        channel_id,
                        user.id,
                        admin_rights=types.ChatAdminRights(other=False),
                        rank=''
                    ))
                    print(f'{num} Добавлен (id: {id})')
                    col = col + 1
                    r=r+1
                    z=z-1
                except:
                    print(f'{num} Не добавлен ')
                    r=r+1
            except:
                print(f'{num} Не добавлен ')
                r = r + 1
        else:
            z=0
    str1 = f'Приглашено пользователей: {col}'
    await event.respond(str1)

client.run_until_disconnected()
